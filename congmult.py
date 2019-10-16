from tabulate import tabulate
from openpyxl import Workbook
from progress.bar import Bar

def congruencialMultiply(x, z):
    g = 0
    bandera2 = z
    while(z>0):
        z = z/2
        g += 1
    g = g + 2
    periodo = 0
    bandera = 0
    mod = 2**g
    n = 2**(g-2)
    a = 3
    data = []
    subdata = []

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Periodo'
    ws['B1'] = 'Xi'
    ws['C1'] = 'X'
    ws['D1'] = 'Ri'
    
    bar = Bar('Processing', max=bandera2)
    while(bandera != x and periodo<=bandera2 and periodo!=n ):
    #while():
        if (periodo == 0):
            bandera = x
        x = (a * x) % mod
        ri = x/(mod-1.0)
        subdata.append(periodo)
        subdata.append('X'+str(periodo))
        subdata.append(x)
        subdata.append(ri)
        ws.append(subdata)
        data.append(subdata)
        subdata=[]
        periodo += 1
        bar.next()
    bar.finish()
    print(tabulate(data, headers=['Periodo', 'Xi', 'X', 'Ri']))
    wb.save('resultsn.xlsx')

def main():
    print('Bienvenido: Algoritmo Congruencial Multiplicativo')
    while True:
        option = raw_input('Introduce cualquier numero para continuar, 0 para salir: \n')
        try:
            option = int(option)
            if(option==0): break
            while True:
                x = int(raw_input("Introduce el valor de la semilla (debe ser impar) 'Xi': "))
                if(x%2==1): break
            wished = int(raw_input("Introduce la cantidad de numeros pseudo-aleatorios deseados "))
            congruencialMultiply(x, wished)
        except Exception as e:
            print('Valor invalido')
            #raise e
    print('Adios!')
 
if __name__ == "__main__":
    main()